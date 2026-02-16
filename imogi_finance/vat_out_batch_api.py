# Copyright (c) 2026, PT. Inovasi Terbaik Bangsa and contributors
# For license information, please see license.txt

"""API endpoints for VAT OUT Batch operations."""

from __future__ import annotations

import frappe
from frappe import _
from frappe.utils import now, getdate

from imogi_finance.tax_operations import generate_vat_out_batch_excel


@frappe.whitelist()
def generate_coretax_upload_file(batch_name: str):
	"""Generate Excel file for CoreTax upload.
	
	Args:
		batch_name: VAT OUT Batch name
		
	Returns:
		dict: File URL and status
	"""
	batch = frappe.get_doc("VAT OUT Batch", batch_name)
	batch.check_permission("write")
	
	# Check if already exported
	if batch.exported_on:
		frappe.throw(_("Batch already exported on {0}. Cancel and create new batch if needed.").format(
			frappe.format(batch.exported_on, {"fieldtype": "Datetime"})
		))
	
	# Check if batch has invoices
	invoices = batch.get_batch_invoices()
	if not invoices:
		frappe.throw(_("No invoices found in batch. Please use 'Get Available Invoices' first."))
	
	# Generate file
	file_url = generate_vat_out_batch_excel(
		batch_name,
		include_fp_numbers=False,
		for_upload=True
	)
	
	# Update batch
	batch.coretax_export_file = file_url
	batch.template_version = "DJP 2024"  # Static version
	batch.exported_on = now()
	batch.save(ignore_permissions=True)
	
	return {
		"status": "success",
		"file_url": file_url,
		"invoice_count": len(invoices),
		"message": _("Export file generated successfully with {0} invoices").format(len(invoices))
	}


@frappe.whitelist()
def generate_reconciliation_file(batch_name: str):
	"""Generate reconciliation Excel file with FP numbers.
	
	Args:
		batch_name: VAT OUT Batch name
		
	Returns:
		dict: File URL and status
	"""
	batch = frappe.get_doc("VAT OUT Batch", batch_name)
	batch.check_permission("read")
	
	# Generate file with FP numbers
	file_url = generate_vat_out_batch_excel(
		batch_name,
		include_fp_numbers=True,
		for_upload=False
	)
	
	# Update batch
	batch.reconciliation_file = file_url
	batch.save(ignore_permissions=True)
	
	return {
		"status": "success",
		"file_url": file_url,
		"message": _("Reconciliation file generated successfully")
	}


@frappe.whitelist()
def import_fp_numbers_from_file(batch_name: str, file_url: str):
	"""Import FP numbers from uploaded Excel file.
	
	Matches rows by Group ID and updates Sales Invoices.
	
	Args:
		batch_name: VAT OUT Batch name
		file_url: Uploaded file URL
		
	Returns:
		dict: Import summary with success/failed counts
	"""
	import openpyxl
	from frappe.utils.file_manager import get_file_path
	
	batch = frappe.get_doc("VAT OUT Batch", batch_name)
	batch.check_permission("write")
	
	# Read Excel file
	try:
		file_path = get_file_path(file_url)
		wb = openpyxl.load_workbook(file_path, data_only=True)
		ws = wb.active
	except Exception as e:
		frappe.throw(_("Could not read file: {0}").format(str(e)))
	
	# Parse header row
	headers = {}
	for idx, cell in enumerate(ws[1], start=1):
		if cell.value:
			headers[str(cell.value).lower().strip()] = idx
	
	# Required columns
	required = ["group id", "fp no seri", "fp no faktur", "fp date"]
	missing = [r for r in required if r not in headers]
	if missing:
		frappe.throw(_("Missing required columns: {0}").format(", ".join(missing)))
	
	# Get batch invoices grouped by group_id
	invoices = batch.get_batch_invoices()
	group_invoices = {}
	for inv in invoices:
		gid = inv.out_fp_group_id
		if gid not in group_invoices:
			group_invoices[gid] = []
		group_invoices[gid].append(inv)
	
	success_count = 0
	failed_count = 0
	warnings = []
	
	# Process data rows
	for row_idx in range(2, ws.max_row + 1):
		try:
			row = ws[row_idx]
			
			# Extract values
			group_id = row[headers["group id"] - 1].value
			fp_no_seri = row[headers["fp no seri"] - 1].value
			fp_no_faktur = row[headers["fp no faktur"] - 1].value
			fp_date = row[headers["fp date"] - 1].value
			
			if not group_id or not fp_no_seri or not fp_no_faktur or not fp_date:
				failed_count += 1
				warnings.append(f"Row {row_idx}: Missing required fields")
				continue
			
			# Convert group_id to int
			try:
				group_id = int(group_id)
			except:
				failed_count += 1
				warnings.append(f"Row {row_idx}: Invalid Group ID: {group_id}")
				continue
			
			# Validate FP number format
			fp_no_clean = str(fp_no_faktur).replace("-", "").replace(".", "").replace(" ", "")
			if len(fp_no_clean) != 16 or not fp_no_clean.isdigit():
				failed_count += 1
				warnings.append(f"Row {row_idx}: Invalid FP Number format: {fp_no_faktur}")
				continue
			
			# Build full FP number
			fp_no_full = f"{fp_no_seri}-{fp_no_faktur}"
			
			# Parse date
			if isinstance(fp_date, str):
				fp_date = getdate(fp_date)
			
			# Find invoices in this group
			if group_id not in group_invoices:
				failed_count += 1
				warnings.append(f"Row {row_idx}: No invoices found for Group {group_id}")
				continue
			
			# Update all invoices in group with same FP number
			for inv in group_invoices[group_id]:
				frappe.db.set_value(
					"Sales Invoice",
					inv.name,
					{
						"out_fp_no": fp_no_full,
						"out_fp_no_seri": fp_no_seri,
						"out_fp_no_faktur": fp_no_faktur,
						"out_fp_date": fp_date
					},
					update_modified=False
				)
			
			success_count += 1
			
		except Exception as e:
			failed_count += 1
			warnings.append(f"Row {row_idx}: {str(e)}")
	
	frappe.db.commit()
	
	# Add comment to batch
	if success_count > 0:
		batch.add_comment("Info", _("Imported {0} FP numbers").format(success_count))
	
	return {
		"status": "success" if failed_count == 0 else "partial",
		"imported_count": success_count,
		"failed_count": failed_count,
		"warnings": warnings[:10],  # Limit warnings
		"message": _("Imported {0} FP numbers, {1} failed").format(success_count, failed_count)
	}
